"""
Lectura de una nota desde un archivo Excel (.xlsx) para importarla al sistema.

El mapa de celdas es el mismo que usa `excel_exports.build_note_like_cell_writes`,
de modo que un Excel generado/llenado con la plantilla de nota de Crystal
(`templates/excel/Nota.xlsx`) se puede volver a importar sin perder datos.

Mapa (hoja 0):
    B9  Nombre        B10 Domicilio    B11 Colonia    B12 Referencias
    I9  F.nacimiento  I10 Telefono     I12 F.elab.    I15 Folio
    B15 F.entrega     B16 F.evento     B17 F.devol.   B43 Instrucciones
    J42 Subtotal  J43 Flete  J44 IVA  J45 Deposito  J46 Total  J47 Desc.  J48 Anticipo
    Articulos filas 21-41: B=cantidad C=equipo I=precio_unit J=total
"""

from __future__ import annotations

import re
import warnings
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

ITEM_START_ROW = 21
ITEM_END_ROW = 41  # inclusive (21 renglones)

# Folio tipo "0001-26" / "45-26".
FOLIO_PATTERN = re.compile(r"^\d{1,4}-\d{2}$")


class NoteImportError(ValueError):
    """Error legible para el usuario al leer un Excel de nota."""


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def read_note_excel(file_obj) -> dict:
    """Lee `file_obj` (.xlsx) y devuelve folio + nota lista para apply_note_to_quotation.

    Lanza NoteImportError con un mensaje legible si el archivo no se puede leer.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - openpyxl es dependencia base
        raise NoteImportError(
            "No se pudo leer el Excel: falta la libreria openpyxl en el servidor.",
        ) from error

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(file_obj, data_only=True)
    except Exception as error:  # noqa: BLE001 - openpyxl lanza varios tipos
        raise NoteImportError(
            "El archivo no es un Excel (.xlsx) valido o esta danado. "
            "Usa la plantilla de nota de Crystal y guardala como .xlsx.",
        ) from error

    worksheet = workbook.worksheets[0]

    def cell(reference: str):
        return worksheet[reference].value

    folio = _to_text(cell("I15"))
    if not FOLIO_PATTERN.fullmatch(folio):
        folio = ""  # se generara uno automaticamente

    client_name = _to_text(cell("B9"))
    address = _to_text(cell("B10"))
    neighborhood = _to_text(cell("B11"))
    reference = _to_text(cell("B12"))
    phone = _to_text(cell("I10"))
    birth_date = _to_date(cell("I9"))
    note_date = _to_date(cell("I12"))

    delivery_date = _to_date(cell("B15"))
    event_date = _to_date(cell("B16"))
    collection_date = _to_date(cell("B17"))
    delivery_instructions = _to_text(cell("B43"))

    subtotal = _to_decimal(cell("J42"))
    freight = _to_decimal(cell("J43"))
    tax_amount = _to_decimal(cell("J44"))
    security_deposit = _to_decimal(cell("J45"))
    total_estimated = _to_decimal(cell("J46"))
    discount = _to_decimal(cell("J47"))
    advance_payment = _to_decimal(cell("J48"))

    equipment_items: list[dict] = []
    for row in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        quantity = _to_int(worksheet.cell(row=row, column=2).value)   # B
        equipment = _to_text(worksheet.cell(row=row, column=3).value)  # C
        if not quantity and not equipment:
            continue
        equipment_items.append(
            {
                "quantity": quantity,
                "equipment": equipment,
                "unitPrice": _to_decimal(worksheet.cell(row=row, column=9).value),   # I
                "total": _to_decimal(worksheet.cell(row=row, column=10).value),      # J
            }
        )

    # Respaldos: si los totales vienen vacios, derivarlos de los articulos.
    items_subtotal = sum((item["total"] for item in equipment_items), Decimal("0"))
    if subtotal <= 0 and items_subtotal > 0:
        subtotal = items_subtotal
    if total_estimated <= 0:
        total_estimated = subtotal + freight + tax_amount + security_deposit

    balance_due = total_estimated - discount - advance_payment

    note = {
        "clientInfo": {
            "fullName": client_name,
            "phoneNumber": phone,
            "birthDate": birth_date,
            "address": address,
            "neighborhood": neighborhood,
            "reference": reference,
            "deliveryInstructions": delivery_instructions,
        },
        "schedule": {
            "deliveryDate": delivery_date,
            "eventDate": event_date,
            "collectionDate": collection_date,
        },
        "logistics": {
            "freight": freight,
            "securityDeposit": security_deposit,
            "applyTax": tax_amount > 0,
        },
        "summary": {
            "subtotal": subtotal,
            "freight": freight,
            "taxAmount": tax_amount,
            "securityDeposit": security_deposit,
            "discount": discount,
            "advancePayment": advance_payment,
            "totalEstimated": total_estimated,
            "balanceDue": balance_due,
        },
        "equipmentItems": equipment_items,
    }

    return {
        "folio": folio,
        "noteDate": note_date,
        "eventDate": event_date,
        "note": note,
    }
