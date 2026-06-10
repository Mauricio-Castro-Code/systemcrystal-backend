from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import os
import platform
import re
import shutil
import subprocess
import tempfile
from typing import Literal
import warnings
import zipfile

from django.conf import settings
from django.utils import timezone

from .models import Order, Quotation


MAX_TEMPLATE_ITEMS = 21
ITEM_START_ROW = 21
CELL_REFERENCE_PATTERN = re.compile(r"([A-Z]+)(\d+)$")
DOCUMENT_PRINT_AREA = "$A$1:$K$72"
DOCUMENT_PAGE_MARGIN_CM = 0.635
DOCUMENT_PAGE_HEADER_MARGIN_CM = 0.423
DOCUMENT_PAGE_FOOTER_MARGIN_CM = 0.423


class ExcelTemplateExportError(ValueError):
    pass


DocumentCellKind = Literal["blank", "text", "number", "date", "formula"]


@dataclass(frozen=True)
class DocumentCellWrite:
    cell_reference: str
    kind: DocumentCellKind
    value: str | int | float | Decimal | date | None


@dataclass(frozen=True)
class GeneratedDocumentBundle:
    excel_bytes: bytes
    excel_filename: str
    pdf_bytes: bytes
    pdf_filename: str


# ── Public API (unchanged signatures) ──────────────────────────────────


def export_quotation_excel(quotation: Quotation) -> tuple[bytes, str]:
    document = render_quotation_documents(quotation)
    return document.excel_bytes, document.excel_filename


def export_quotation_pdf(quotation: Quotation) -> tuple[bytes, str]:
    document = render_quotation_documents(quotation)
    return document.pdf_bytes, document.pdf_filename


def export_order_excel(order: Order) -> tuple[bytes, str]:
    document = render_order_documents(order)
    return document.excel_bytes, document.excel_filename


def export_order_pdf(order: Order) -> tuple[bytes, str]:
    document = render_order_documents(order)
    return document.pdf_bytes, document.pdf_filename


def render_quotation_documents(quotation: Quotation) -> GeneratedDocumentBundle:
    return render_note_like_documents(
        document_id=quotation.quotation_id,
        created_at=timezone.localtime(quotation.created_at),
        quotation=quotation,
        template_path=settings.QUOTATION_EXCEL_TEMPLATE_PATH,
    )


def render_order_documents(order: Order) -> GeneratedDocumentBundle:
    return render_note_like_documents(
        document_id=order.order_id,
        created_at=timezone.localtime(order.confirmed_at),
        quotation=order.quotation,
        template_path=settings.NOTE_EXCEL_TEMPLATE_PATH,
    )


# ── Core rendering pipeline ────────────────────────────────────────────


def render_note_like_documents(
    *,
    document_id: str,
    created_at: datetime,
    quotation: Quotation,
    template_path: Path,
) -> GeneratedDocumentBundle:
    normalized_template_path = Path(template_path)

    if not normalized_template_path.exists():
        raise FileNotFoundError(
            f"No se encontro la plantilla Excel en {normalized_template_path}.",
        )

    xlsx_template_path = resolve_xlsx_template(normalized_template_path)

    cell_writes = build_note_like_cell_writes(
        document_id=document_id,
        created_at=created_at,
        quotation=quotation,
    )

    return render_document_bundle(
        template_path=xlsx_template_path,
        output_stem=document_id,
        cell_writes=cell_writes,
    )


def build_note_like_cell_writes(
    *,
    document_id: str,
    created_at: datetime,
    quotation: Quotation,
) -> list[DocumentCellWrite]:
    equipment_items = list(quotation.equipment_items.all())

    if len(equipment_items) > MAX_TEMPLATE_ITEMS:
        raise ExcelTemplateExportError(
            "La plantilla solo soporta hasta 21 renglones de equipo.",
        )

    writes: list[DocumentCellWrite] = [
        text_cell("B9", quotation.client_name),
        text_cell("B10", quotation.address),
        text_cell("B11", quotation.neighborhood),
        text_cell("B12", quotation.reference),
        date_cell("I9", quotation.birth_date),
        text_cell("I10", quotation.phone_number),
        date_cell("I12", created_at.date()),
        text_cell("I15", document_id),
        date_cell("B15", quotation.delivery_date),
        date_cell("B16", quotation.event_date),
        date_cell("B17", quotation.collection_date),
        text_cell("B43", quotation.delivery_instructions),
        number_cell("J42", quotation.subtotal),
        number_cell("J43", quotation.freight),
        number_cell("J44", quotation.tax_amount),
        number_cell("J45", quotation.security_deposit),
        number_cell("J46", quotation.total_estimated),
        number_cell("J47", quotation.discount),
        number_cell("J48", quotation.advance_payment),
        formula_cell("J49", "=J46-J47-J48"),
    ]

    for row_index in range(MAX_TEMPLATE_ITEMS):
        row_number = ITEM_START_ROW + row_index
        item = equipment_items[row_index] if row_index < len(equipment_items) else None

        writes.extend(
            [
                number_cell(f"B{row_number}", item.quantity if item else None),
                text_cell(f"C{row_number}", item.equipment if item else None),
                number_cell(f"I{row_number}", item.unit_price if item else None),
                number_cell(f"J{row_number}", item.total if item else None),
            ],
        )

    return writes


# ── Template format detection & conversion ────────────────────────────


def _is_legacy_xls(path: Path) -> bool:
    """Return True if the file is old binary .xls format (not a ZIP-based .xlsx)."""
    try:
        with zipfile.ZipFile(path):
            return False
    except (zipfile.BadZipFile, Exception):
        return True


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    """
    Convert a legacy .xls template to .xlsx using LibreOffice.
    The converted file is placed next to the original with a .xlsx suffix.
    Raises ExcelTemplateExportError if conversion fails.
    """
    soffice = _resolve_soffice_binary()

    if not soffice:
        raise ExcelTemplateExportError(
            "La plantilla esta en formato .xls antiguo y no hay LibreOffice disponible "
            "para convertirla. Abre el archivo en Excel y guardalo como .xlsx."
        )

    out_dir = xls_path.parent
    stem = xls_path.stem

    try:
        result = subprocess.run(
            [
                soffice,
                "--headless",
                "--norestore",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(out_dir),
                str(xls_path),
            ],
            capture_output=True,
            timeout=60,
            check=False,
        )
    except (FileNotFoundError, OSError, subprocess.TimeoutExpired) as exc:
        raise ExcelTemplateExportError(
            f"Error al convertir la plantilla con LibreOffice: {exc}"
        ) from exc

    if result.returncode != 0:
        raise ExcelTemplateExportError(
            "LibreOffice no pudo convertir la plantilla al formato .xlsx."
        )

    converted = out_dir / f"{stem}.xlsx"

    if not converted.exists():
        raise ExcelTemplateExportError(
            "LibreOffice termino pero no genero el archivo .xlsx esperado."
        )

    return converted


def resolve_xlsx_template(template_path: Path) -> Path:
    """
    Return a valid .xlsx path for the given template.
    If the template is already proper .xlsx, return it as-is.
    If it is a legacy .xls binary file, convert it and return the converted path.
    """
    if not _is_legacy_xls(template_path):
        return template_path

    converted = template_path.with_suffix(".xlsx")
    if converted != template_path and converted.exists():
        if converted.stat().st_mtime >= template_path.stat().st_mtime:
            return converted

    return _convert_xls_to_xlsx(template_path)


# ── openpyxl-based rendering ───────────────────────────────────────────


def render_document_bundle(
    *,
    template_path: Path,
    output_stem: str,
    cell_writes: list[DocumentCellWrite],
) -> GeneratedDocumentBundle:
    from openpyxl import load_workbook
    from openpyxl.worksheet.page import PageMargins, PrintPageSetup

    with tempfile.TemporaryDirectory(prefix="crystal-documents-") as tmp:
        tmp_path = Path(tmp)
        excel_output_path = tmp_path / f"{output_stem}.xlsx"
        pdf_output_path = tmp_path / f"{output_stem}.pdf"

        # openpyxl emite warnings ruidosos por imagenes WMF de la plantilla.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(str(template_path))
        ws = wb.worksheets[0]

        _apply_page_setup(ws)
        _apply_cell_writes(ws, cell_writes)

        wb.save(str(excel_output_path))
        wb.close()

        if not excel_output_path.exists():
            raise ExcelTemplateExportError(
                "openpyxl no genero el archivo Excel esperado.",
            )

        # openpyxl descarta imagenes WMF (logos, iconos) al guardar.
        # Las copiamos a nivel ZIP desde la plantilla original.
        _restore_xlsx_media(template_path, excel_output_path)

        _generate_pdf(excel_output_path, pdf_output_path, document_id=output_stem, cell_writes=cell_writes)

        if not pdf_output_path.exists():
            raise ExcelTemplateExportError(
                "No fue posible generar la vista PDF del documento.",
            )

        return GeneratedDocumentBundle(
            excel_bytes=excel_output_path.read_bytes(),
            excel_filename=excel_output_path.name,
            pdf_bytes=pdf_output_path.read_bytes(),
            pdf_filename=pdf_output_path.name,
        )


def _apply_page_setup(ws) -> None:
    ws.print_area = DOCUMENT_PRINT_AREA
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 1  # Letter
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.page_margins = ws.page_margins or type(ws.page_margins)()
    ws.page_margins.left = DOCUMENT_PAGE_MARGIN_CM
    ws.page_margins.right = DOCUMENT_PAGE_MARGIN_CM
    ws.page_margins.top = DOCUMENT_PAGE_MARGIN_CM
    ws.page_margins.bottom = DOCUMENT_PAGE_MARGIN_CM
    ws.page_margins.header = DOCUMENT_PAGE_HEADER_MARGIN_CM
    ws.page_margins.footer = DOCUMENT_PAGE_FOOTER_MARGIN_CM

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False


def _apply_cell_writes(ws, cell_writes: list[DocumentCellWrite]) -> None:
    from openpyxl.styles import PatternFill

    # LibreOffice renderiza GRIS las celdas escritas por openpyxl que no
    # tienen fill explicito. Forzamos fondo blanco en cada escritura.
    blanco = PatternFill(fill_type="solid", fgColor="FFFFFFFF")

    for cw in cell_writes:
        cell = ws[cw.cell_reference]

        if cw.kind == "blank":
            cell.value = None
        elif cw.kind == "text":
            cell.value = str(cw.value or "")
        elif cw.kind == "number":
            cell.value = _to_float(cw.value)
        elif cw.kind == "date":
            cell.value = cw.value
        elif cw.kind == "formula":
            cell.value = str(cw.value or "")

        if cw.kind != "blank":
            existing_fill = cell.fill
            fill_type = getattr(existing_fill, "fill_type", None)
            if fill_type in (None, "none"):
                cell.fill = blanco


# Carpetas del .xlsx que openpyxl NO entiende y descarta silenciosamente al
# guardar. Las copiamos enteras desde la plantilla original al output para
# que el archivo final conserve logos WMF/EMF, drawings, textboxes con
# clausulas, pagares OLE, charts, vmlDrawings, etc.
#
# Si el comando `python manage.py diagnose_excel_template` reporta otra
# carpeta, agregala aqui.
_TEMPLATE_PRESERVED_PREFIXES = (
    "xl/media/",            # imagenes (WMF, EMF, PNG, JPEG)
    "xl/drawings/",         # shapes, textboxes, anclas de imagenes
    "xl/embeddings/",       # objetos OLE (pagare, contratos PDF)
    "xl/charts/",           # graficos
    "xl/printerSettings/",  # config de impresora
    "xl/theme/",            # tema custom corporativo
    "xl/worksheets/_rels/", # relaciones hoja->drawings/oleObjects
    "xl/_rels/",            # relaciones del workbook
)


_SHEET_DATA_PATTERN = re.compile(
    r"<sheetData(?:\s[^>]*)?(?:/>|>.*?</sheetData>)",
    re.DOTALL,
)


def _restore_xlsx_media(template_path: Path, output_path: Path) -> None:
    """
    Reconstruye el .xlsx final usando el TEMPLATE como base y trasplantando
    del output de openpyxl solo lo estrictamente necesario:
      - <sheetData> de cada hoja  (las celdas con valores nuevos)
      - xl/sharedStrings.xml      (textos referenciados por celdas)
      - xl/calcChain.xml          (cache de formulas)

    Por que asi en lugar de "patchear" lo que openpyxl rompe:
      openpyxl no solo descarta archivos (WMF, oleObjects, embeddings); a
      veces preserva el archivo binario PERO elimina del XML de la hoja la
      referencia que lo apunta (<drawing r:id="..."/>, <legacyDrawing/>,
      <oleObjects>, etc.). El resultado: el ZIP tiene la imagen pero la
      hoja ya no la usa, y LibreOffice no la dibuja.

      Tomando el sheet XML del template y solo trasplantando el bloque
      <sheetData>, openpyxl deja de tener oportunidad de tocar las
      referencias visuales.
    """
    try:
        with zipfile.ZipFile(template_path, "r") as zin:
            template_data: dict[str, bytes] = {
                name: zin.read(name) for name in zin.namelist()
            }
        with zipfile.ZipFile(output_path, "r") as zout:
            output_data: dict[str, bytes] = {
                name: zout.read(name) for name in zout.namelist()
            }
    except zipfile.BadZipFile:
        return

    if not template_data:
        return

    # Arrancamos con TODO el template
    final_data: dict[str, bytes] = dict(template_data)

    # Trasplantar shared strings y calcChain del output (datos nuevos).
    for name in ("xl/sharedStrings.xml", "xl/calcChain.xml"):
        if name in output_data:
            final_data[name] = output_data[name]

    # En cada hoja, trasplantar solo el bloque <sheetData> del output
    # y parchear pageSetup para que LibreOffice respete fit-to-one-page.
    for name, output_sheet in output_data.items():
        if not (
            name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ):
            continue
        if name not in template_data:
            final_data[name] = output_sheet
            continue
        merged = _transplant_sheet_data(template_data[name], output_sheet)
        final_data[name] = _patch_page_setup(merged)

    # Si openpyxl creo archivos nuevos no presentes en el template, agregarlos.
    for name, data in output_data.items():
        if name not in final_data:
            final_data[name] = data

    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    try:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zfinal:
            for name, data in final_data.items():
                zfinal.writestr(name, data)
        os.replace(tmp_path, output_path)
    except Exception:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)
        raise


_PAGE_SETUP_PATTERN = re.compile(r"<pageSetup\b[^/]*/?>")
_PAGE_SETUP_PR_PATTERN = re.compile(r"<pageSetUpPr\b[^>]*/?>")

# Escala fija (porcentaje) que ajusta el area de impresion A1:K72 a una sola
# pagina Letter. Medida empiricamente con LibreOffice: 87% es la escala mas
# grande que todavia cabe en una pagina. Usar una escala EXPLICITA (en vez de
# fitToWidth/fitToHeight) garantiza que el PDF se vea identico en cualquier
# version de LibreOffice -macOS local y Linux en el servidor-, porque no
# depende del calculo de "ajustar a pagina" que cada version hace distinto.
_FIXED_PAGE_SCALE = 87


def _patch_page_setup(sheet_bytes: bytes) -> bytes:
    """
    Fija una escala explicita (paperSize Letter + scale=87) y elimina
    fitToPage del pageSetUpPr. Con fitToPage activo LibreOffice ignora el
    atributo scale y recalcula el ajuste, lo que produce tamanos de fuente
    distintos entre macOS y Linux. Con scale explicito el render es identico.
    """
    try:
        text = sheet_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return sheet_bytes

    # Quitar fitToPage del pageSetUpPr para que el scale explicito tenga efecto.
    text = _PAGE_SETUP_PR_PATTERN.sub("", text)

    replacement = (
        '<pageSetup paperSize="1" orientation="portrait" '
        f'scale="{_FIXED_PAGE_SCALE}" '
        'horizontalDpi="300" verticalDpi="300"/>'
    )

    if _PAGE_SETUP_PATTERN.search(text):
        text = _PAGE_SETUP_PATTERN.sub(replacement, text, count=1)
    else:
        # Insert before </worksheet> if no pageSetup tag exists.
        text = text.replace("</worksheet>", replacement + "</worksheet>", 1)

    return text.encode("utf-8")


def _transplant_sheet_data(
    template_sheet_bytes: bytes,
    output_sheet_bytes: bytes,
) -> bytes:
    """
    Reemplaza el bloque <sheetData>...</sheetData> del sheet XML del template
    por el del output (que trae las celdas llenas). Todo lo demas del sheet
    -drawings, oleObjects, mergeCells, hyperlinks, page setup, etc.- se queda
    como estaba en el template.
    """
    try:
        template_text = template_sheet_bytes.decode("utf-8")
        output_text = output_sheet_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return output_sheet_bytes

    output_match = _SHEET_DATA_PATTERN.search(output_text)
    template_match = _SHEET_DATA_PATTERN.search(template_text)

    if not output_match or not template_match:
        return output_sheet_bytes

    new_sheet_data = output_match.group(0)
    merged_text = (
        template_text[: template_match.start()]
        + new_sheet_data
        + template_text[template_match.end():]
    )
    return merged_text.encode("utf-8")


def _to_float(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    return float(Decimal(str(value)))


# ── LibreOffice binary discovery ───────────────────────────────────────


_LIBREOFFICE_KNOWN_PATHS = (
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/local/bin/soffice",
    "/opt/homebrew/bin/soffice",
)


def _resolve_soffice_binary() -> str | None:
    """
    Busca el binario de LibreOffice en este orden:
      1. settings.LIBREOFFICE_BINARY o env LIBREOFFICE_BINARY
      2. PATH (soffice / libreoffice)
      3. Rutas conocidas de macOS y Linux
    """
    configured = (
        getattr(settings, "LIBREOFFICE_BINARY", "")
        or os.getenv("LIBREOFFICE_BINARY", "")
    ).strip()
    if configured and Path(configured).exists():
        return configured

    in_path = shutil.which("soffice") or shutil.which("libreoffice")
    if in_path:
        return in_path

    for candidate in _LIBREOFFICE_KNOWN_PATHS:
        if Path(candidate).exists():
            return candidate

    return None


# ── PDF generation (LibreOffice → native Excel fallback) ───────────────


def _generate_pdf(
    excel_path: Path,
    pdf_path: Path,
    *,
    document_id: str = "",
    cell_writes: "list[DocumentCellWrite] | None" = None,
) -> None:
    if _try_libreoffice_pdf(excel_path, pdf_path):
        return

    if _try_native_excel_pdf(excel_path, pdf_path):
        return

    if cell_writes is not None and _try_fpdf2_pdf(pdf_path, document_id, cell_writes):
        return

    raise ExcelTemplateExportError(
        "No se encontro LibreOffice ni Microsoft Excel para generar PDF. "
        "Instala LibreOffice (soffice) o Microsoft Excel.",
    )


def _try_libreoffice_pdf(excel_path: Path, pdf_path: Path) -> bool:
    soffice = _resolve_soffice_binary()
    if not soffice:
        return False

    # Provide a writable HOME for LibreOffice profile (needed in Docker).
    # Force Spanish locale so date cells render in Spanish.
    env = os.environ.copy()
    env.setdefault("HOME", str(pdf_path.parent))
    env.setdefault("LANG", "es_MX.UTF-8")
    env.setdefault("LC_ALL", "es_MX.UTF-8")

    try:
        result = subprocess.run(
            [
                soffice,
                "--headless",
                "--norestore",
                "--nofirststartwizard",
                "--convert-to",
                "pdf",
                "--outdir",
                str(pdf_path.parent),
                str(excel_path),
            ],
            env=env,
            capture_output=True,
            timeout=getattr(settings, "DOCUMENT_RENDER_TIMEOUT_SECONDS", 120),
            check=False,
        )
    except (FileNotFoundError, OSError, subprocess.TimeoutExpired):
        return False

    if result.returncode != 0:
        return False

    generated = pdf_path.parent / f"{excel_path.stem}.pdf"
    if generated.exists() and generated != pdf_path:
        generated.rename(pdf_path)

    return pdf_path.exists()


def _try_native_excel_pdf(excel_path: Path, pdf_path: Path) -> bool:
    system_name = platform.system()
    timeout = getattr(settings, "DOCUMENT_RENDER_TIMEOUT_SECONDS", 120)

    if system_name == "Darwin":
        return _try_macos_excel_pdf(excel_path, pdf_path, timeout)

    if system_name == "Windows":
        return _try_windows_excel_pdf(excel_path, pdf_path, timeout)

    return False


def _try_macos_excel_pdf(
    excel_path: Path,
    pdf_path: Path,
    timeout: int,
) -> bool:
    script = "\n".join([
        'tell application "Microsoft Excel"',
        "set display alerts to false",
        f'open POSIX file "{excel_path}"',
        "set wb to active workbook",
        f'save workbook as wb filename "{pdf_path}" file format PDF file format',
        "close wb saving no",
        "end tell",
    ])

    try:
        result = subprocess.run(
            ["osascript", "-"],
            input=script,
            text=True,
            capture_output=True,
            timeout=timeout,
            check=False,
        )
        return result.returncode == 0 and pdf_path.exists()
    except (FileNotFoundError, OSError, subprocess.TimeoutExpired):
        return False


def _try_windows_excel_pdf(
    excel_path: Path,
    pdf_path: Path,
    timeout: int,
) -> bool:
    ps_script = "\n".join([
        "$ErrorActionPreference = 'Stop'",
        "$excel = $null",
        "$wb = $null",
        "try {",
        "  $excel = New-Object -ComObject Excel.Application",
        "  $excel.Visible = $false",
        "  $excel.DisplayAlerts = $false",
        f"  $wb = $excel.Workbooks.Open('{excel_path}')",
        f"  $wb.ExportAsFixedFormat(0, '{pdf_path}')",
        "  $wb.Close($false)",
        "} finally {",
        "  if ($wb) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }",
        "  if ($excel) { $excel.Quit(); [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }",
        "  [GC]::Collect()",
        "}",
    ])

    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", suffix=".ps1", delete=False,
    ) as f:
        f.write(ps_script)
        ps_path = Path(f.name)

    try:
        result = subprocess.run(
            [
                "powershell.exe", "-NoProfile", "-NonInteractive",
                "-ExecutionPolicy", "Bypass", "-File", str(ps_path),
            ],
            capture_output=True,
            timeout=timeout,
            check=False,
        )
        return result.returncode == 0 and pdf_path.exists()
    except (FileNotFoundError, OSError, subprocess.TimeoutExpired):
        return False
    finally:
        ps_path.unlink(missing_ok=True)


# ── fpdf2 fallback ────────────────────────────────────────────────────


def _try_fpdf2_pdf(
    pdf_path: Path,
    document_id: str,
    cell_writes: "list[DocumentCellWrite]",
) -> bool:
    """Genera un PDF con fpdf2 a partir de cell_writes cuando LibreOffice no esta disponible."""
    try:
        from fpdf import FPDF  # type: ignore[import]

        data = {cw.cell_reference: cw.value for cw in cell_writes}

        def _currency(val) -> str:
            try:
                return f"${float(val):,.2f}" if val is not None else ""
            except (TypeError, ValueError):
                return str(val) if val else ""

        def _date(val) -> str:
            if val is None:
                return ""
            if hasattr(val, "strftime"):
                return val.strftime("%d/%m/%Y")
            return str(val)

        def _txt(val) -> str:
            return str(val).strip() if val is not None else ""

        PAGE_W = 183.0  # usable width: Letter 215.9 - 2*16.5 margins

        pdf = FPDF(orientation="P", unit="mm", format="Letter")
        pdf.set_margins(16.5, 14, 16.5)
        pdf.set_auto_page_break(auto=True, margin=14)
        pdf.add_page()

        # ── Encabezado ──
        pdf.set_font("Helvetica", "B", 17)
        pdf.cell(PAGE_W, 9, "Crystal Alquiler", ln=True, align="C")
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(PAGE_W, 5, "Renta de Vajilla, Sillas, Mesas y Loza para Eventos", ln=True, align="C")
        pdf.ln(2)
        pdf.set_draw_color(29, 47, 88)
        pdf.set_line_width(0.5)
        pdf.line(16.5, pdf.get_y(), 199.4, pdf.get_y())
        pdf.ln(4)

        # ── Folio y fecha ──
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(PAGE_W * 0.6, 7, f"Nota de Pedido: {document_id}")
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(PAGE_W * 0.4, 7, f"Fecha: {_date(data.get('I12'))}", ln=True, align="R")
        pdf.ln(3)

        def _section_header(title: str) -> None:
            pdf.set_fill_color(29, 47, 88)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(PAGE_W, 6, f"  {title}", ln=True, fill=True)
            pdf.set_text_color(0, 0, 0)
            pdf.ln(1)

        def _info_row(label: str, value: str) -> None:
            if not value:
                return
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(42, 5, f"{label}:")
            pdf.set_font("Helvetica", "", 8)
            pdf.cell(PAGE_W - 42, 5, value, ln=True)

        # ── Datos del cliente ──
        _section_header("DATOS DEL CLIENTE")
        _info_row("Cliente", _txt(data.get("B9")))
        _info_row("Direccion", _txt(data.get("B10")))
        _info_row("Colonia", _txt(data.get("B11")))
        _info_row("Referencia", _txt(data.get("B12")))
        _info_row("Telefono", _txt(data.get("I10")))
        _info_row("Fecha de nacimiento", _date(data.get("I9")))
        pdf.ln(3)

        # ── Fechas ──
        _section_header("FECHAS DE SERVICIO")
        _info_row("Entrega", _date(data.get("B15")))
        _info_row("Evento", _date(data.get("B16")))
        _info_row("Recoleccion", _date(data.get("B17")))
        pdf.ln(3)

        # ── Tabla de equipo ──
        _section_header("EQUIPO RENTADO")

        COL_QTY, COL_DESC, COL_UNIT, COL_TOTAL = 16, 107, 30, 30

        pdf.set_fill_color(235, 239, 246)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(COL_QTY, 6, "Cant.", border=1, fill=True, align="C")
        pdf.cell(COL_DESC, 6, "Descripcion", border=1, fill=True)
        pdf.cell(COL_UNIT, 6, "P. Unitario", border=1, fill=True, align="R")
        pdf.cell(COL_TOTAL, 6, "Total", border=1, fill=True, align="R", ln=True)

        pdf.set_font("Helvetica", "", 8)
        has_items = False
        for i in range(MAX_TEMPLATE_ITEMS):
            row = ITEM_START_ROW + i
            qty = data.get(f"B{row}")
            desc = data.get(f"C{row}")
            unit_price = data.get(f"I{row}")
            total = data.get(f"J{row}")
            if qty is None and not desc:
                continue
            has_items = True
            fill_row = i % 2 == 1
            if fill_row:
                pdf.set_fill_color(249, 250, 252)
            qty_str = str(int(float(qty))) if qty is not None else ""
            pdf.cell(COL_QTY, 5.5, qty_str, border=1, fill=fill_row, align="C")
            pdf.cell(COL_DESC, 5.5, _txt(desc), border=1, fill=fill_row)
            pdf.cell(COL_UNIT, 5.5, _currency(unit_price), border=1, fill=fill_row, align="R")
            pdf.cell(COL_TOTAL, 5.5, _currency(total), border=1, fill=fill_row, align="R", ln=True)

        if not has_items:
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(PAGE_W, 5.5, "Sin articulos registrados.", border=1, ln=True, align="C")

        pdf.ln(4)

        # ── Resumen financiero ──
        _section_header("RESUMEN FINANCIERO")

        try:
            balance = float(data.get("J46") or 0) - float(data.get("J47") or 0) - float(data.get("J48") or 0)
        except (TypeError, ValueError):
            balance = 0.0

        fin_rows = [
            ("Subtotal", data.get("J42")),
            ("Flete", data.get("J43")),
            ("IVA (16%)", data.get("J44")),
            ("Deposito de seguridad", data.get("J45")),
            ("Total estimado", data.get("J46")),
            ("Descuento", data.get("J47")),
            ("Anticipo", data.get("J48")),
        ]
        INDENT = PAGE_W - 95
        pdf.set_font("Helvetica", "", 9)
        for label, val in fin_rows:
            if val is not None:
                try:
                    if float(val) == 0:
                        continue
                except (TypeError, ValueError):
                    pass
            else:
                continue
            pdf.cell(INDENT, 5.5, "")
            pdf.cell(55, 5.5, f"{label}:")
            pdf.cell(40, 5.5, _currency(val), ln=True, align="R")

        pdf.set_fill_color(29, 47, 88)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(INDENT, 7, "", fill=True)
        pdf.cell(55, 7, "Saldo por pagar:", fill=True)
        pdf.cell(40, 7, _currency(balance), ln=True, align="R", fill=True)
        pdf.set_text_color(0, 0, 0)

        # ── Instrucciones ──
        instructions = _txt(data.get("B43"))
        if instructions:
            pdf.ln(4)
            _section_header("INSTRUCCIONES DE ENTREGA")
            pdf.set_font("Helvetica", "", 8)
            pdf.multi_cell(PAGE_W, 5, instructions)

        # ── Footer ──
        pdf.ln(6)
        pdf.set_line_width(0.4)
        pdf.line(16.5, pdf.get_y(), 199.4, pdf.get_y())
        pdf.ln(2)
        pdf.set_font("Helvetica", "I", 7)
        pdf.set_text_color(130, 130, 130)
        pdf.cell(PAGE_W, 4, "Crystal Alquiler - Documento generado electronicamente", ln=True, align="C")

        pdf.output(str(pdf_path))
        return pdf_path.exists()

    except Exception:
        return False


# ── Cell write helpers (unchanged) ─────────────────────────────────────


def text_cell(cell_reference: str, value) -> DocumentCellWrite:
    normalized_reference = validate_cell_reference(cell_reference)
    normalized_text = normalize_document_text(value)

    if not normalized_text:
        return blank_cell(normalized_reference)

    return DocumentCellWrite(normalized_reference, "text", normalized_text)


def number_cell(cell_reference: str, value) -> DocumentCellWrite:
    normalized_reference = validate_cell_reference(cell_reference)

    if value is None or value == "":
        return blank_cell(normalized_reference)

    if isinstance(value, Decimal):
        return DocumentCellWrite(normalized_reference, "number", value)

    if isinstance(value, int | float):
        return DocumentCellWrite(normalized_reference, "number", value)

    try:
        return DocumentCellWrite(normalized_reference, "number", Decimal(str(value)))
    except Exception as error:  # noqa: BLE001
        raise ExcelTemplateExportError(
            f"La celda {cell_reference} requiere un numero valido.",
        ) from error


def date_cell(cell_reference: str, value) -> DocumentCellWrite:
    normalized_reference = validate_cell_reference(cell_reference)

    if value in {None, ""}:
        return blank_cell(normalized_reference)

    if isinstance(value, datetime):
        return DocumentCellWrite(normalized_reference, "date", value.date())

    if isinstance(value, date):
        return DocumentCellWrite(normalized_reference, "date", value)

    raise ExcelTemplateExportError(
        f"La celda {cell_reference} requiere una fecha valida.",
    )


def formula_cell(cell_reference: str, value: str) -> DocumentCellWrite:
    normalized_reference = validate_cell_reference(cell_reference)
    normalized_formula = str(value or "").strip()

    if not normalized_formula:
        return blank_cell(normalized_reference)

    return DocumentCellWrite(normalized_reference, "formula", normalized_formula)


def blank_cell(cell_reference: str) -> DocumentCellWrite:
    return DocumentCellWrite(validate_cell_reference(cell_reference), "blank", None)


def validate_cell_reference(cell_reference: str) -> str:
    normalized_reference = str(cell_reference or "").strip().upper()

    if not CELL_REFERENCE_PATTERN.fullmatch(normalized_reference):
        raise ExcelTemplateExportError(
            f"La referencia de celda {cell_reference!r} no es valida.",
        )

    return normalized_reference


def normalize_document_text(value) -> str:
    normalized_text = str(value or "").strip()
    return normalized_text.upper()
